"""
Transport Chain Control System
============================
Main entry point for the system
"""

import yaml
import argparse
from pathlib import Path
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# models 导入
from models import (
    create_models,
    get_default_params,
    SystemModel,
    JacobianCalculator
)

# RL 导入
from RL.config import TrainingConfig, SystemConfig
from RL.experiment import create_rl_system, train_agent, evaluate_agent
from RL.utils.helpers import set_random_seed
from RL.utils.logger import Logger
from RL.utils.plotter import Plotter
from RL.exceptions import RLError

# bifurcation 导入
from bifurcation import (
    BifurcationAnalyzer,
    BifurcationParameters,
    DEFAULT_PARAMS
)

def save_timestep_data_to_excel(
        filename, rail_chains, sea_chains, 
        q_rail_history, q_sea_history,
        price_rail_history, price_sea_history,
        simulation_time):
    """Save each time step's data to a separate sheet in an Excel file."""
    
    wb = Workbook()
    
    # Create header row
    headers = ["Chain Type", "Chain Name", "Company Index", "Flow", "Price", "Total Chain Flow"]
    
    # For each time step
    for t_index, t in enumerate(simulation_time):
        # Create new sheet for this time step
        sheet_name = f"Time_{t:.2f}"
        if t_index == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)
        
        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        current_row = 2
        
        # Write rail chain data
        for chain_index, chain in enumerate(rail_chains):
            chain_flows = q_rail_history[t_index][chain_index]
            chain_prices = price_rail_history[t_index][chain_index]
            total_chain_flow = sum(chain_flows)
            
            for company_index, (flow, price) in enumerate(zip(chain_flows, chain_prices)):
                ws.cell(row=current_row, column=1, value="Rail")
                ws.cell(row=current_row, column=2, value=chain.name)
                ws.cell(row=current_row, column=3, value=company_index)
                ws.cell(row=current_row, column=4, value=flow)
                ws.cell(row=current_row, column=5, value=price)
                ws.cell(row=current_row, column=6, value=total_chain_flow)
                current_row += 1
            
            current_row += 1
        
        # Write sea chain data
        for chain_index, chain in enumerate(sea_chains):
            chain_flows = q_sea_history[t_index][chain_index]
            chain_prices = price_sea_history[t_index][chain_index]
            total_chain_flow = sum(chain_flows)
            
            for company_index, (flow, price) in enumerate(zip(chain_flows, chain_prices)):
                ws.cell(row=current_row, column=1, value="Sea")
                ws.cell(row=current_row, column=2, value=chain.name)
                ws.cell(row=current_row, column=3, value=company_index)
                ws.cell(row=current_row, column=4, value=flow)
                ws.cell(row=current_row, column=5, value=price)
                ws.cell(row=current_row, column=6, value=total_chain_flow)
                current_row += 1
            
            current_row += 1
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Add summary sheet
    summary_sheet = wb.create_sheet("Summary", 0)
    summary_headers = ["Time Step", "Total Rail Flow", "Total Sea Flow", 
                      "Average Rail Price", "Average Sea Price"]
    
    for col, header in enumerate(summary_headers, 1):
        summary_sheet.cell(row=1, column=col, value=header)
    
    for t_index, t in enumerate(simulation_time, 2):
        total_rail_flow = sum(sum(chain_flows) for chain_flows in q_rail_history[t_index-2])
        total_sea_flow = sum(sum(chain_flows) for chain_flows in q_sea_history[t_index-2])
        avg_rail_price = sum(sum(prices) for prices in price_rail_history[t_index-2]) / \
                        sum(len(prices) for prices in price_rail_history[t_index-2])
        avg_sea_price = sum(sum(prices) for prices in price_sea_history[t_index-2]) / \
                       sum(len(prices) for prices in price_sea_history[t_index-2])
        
        summary_sheet.cell(row=t_index, column=1, value=t)
        summary_sheet.cell(row=t_index, column=2, value=total_rail_flow)
        summary_sheet.cell(row=t_index, column=3, value=total_sea_flow)
        summary_sheet.cell(row=t_index, column=4, value=avg_rail_price)
        summary_sheet.cell(row=t_index, column=5, value=avg_sea_price)
    
    for col in range(1, len(summary_headers) + 1):
        summary_sheet.column_dimensions[get_column_letter(col)].width = 15
    
    wb.save(filename)
    print(f"\nData saved to {filename}")
    print(f"Total sheets: {len(wb.sheetnames)}")

def analyze_and_save_jacobian(jacobian_history, output_path):
    """Analyze and save Jacobian matrix results with improved numerical stability"""
    wb = Workbook()
    
    ws_ode = wb.create_sheet("ODE Jacobian Analysis")
    ws_pde = wb.create_sheet("PDE Jacobian Analysis")
    wb.remove(wb['Sheet'])
    
    for ws in [ws_ode, ws_pde]:
        for col in range(1, 20):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    ws_ode.append(["Time", "Max Eigenvalue", "Min Eigenvalue", "Condition Number", 
                   "Trace", "Log Determinant", "Stability"])
    ws_pde.append(["Time", "Max Eigenvalue", "Min Eigenvalue", "Condition Number", 
                   "Trace", "Log Determinant", "Stability"])
    
    ws_details = wb.create_sheet("Detailed Matrices")
    current_row = 1
    
    for entry in jacobian_history:
        time = entry['time']
        
        try:
            ode_jac = entry['ode_jacobian']
            ode_eigenvals = np.linalg.eigvals(ode_jac)
            ode_max_eigen = np.max(np.real(ode_eigenvals))
            ode_min_eigen = np.min(np.real(ode_eigenvals))
            
            _, s, _ = np.linalg.svd(ode_jac)
            ode_cond = s[0] / s[-1] if len(s) > 0 else np.inf
            
            ode_trace = np.trace(ode_jac)
            
            sign, logdet = np.linalg.slogdet(ode_jac)
            ode_logdet = logdet if sign >= 0 else -logdet
            
            ode_stability = "Stable" if np.all(np.real(ode_eigenvals) < 0) else "Unstable"
            
            ws_ode.append([
                time, ode_max_eigen, ode_min_eigen, ode_cond, 
                ode_trace, ode_logdet, ode_stability
            ])
        except Exception as e:
            print(f"Error analyzing ODE Jacobian at time {time}: {str(e)}")
            ws_ode.append([time, "Error", "Error", "Error", "Error", "Error", "Error"])
        
        try:
            pde_jac = entry['pde_jacobian']
            pde_eigenvals = np.linalg.eigvals(pde_jac)
            pde_max_eigen = np.max(np.real(pde_eigenvals))
            pde_min_eigen = np.min(np.real(pde_eigenvals))
            
            _, s, _ = np.linalg.svd(pde_jac)
            pde_cond = s[0] / s[-1] if len(s) > 0 else np.inf
            
            pde_trace = np.trace(pde_jac)
            
            sign, logdet = np.linalg.slogdet(pde_jac)
            pde_logdet = logdet if sign >= 0 else -logdet
            
            pde_stability = "Stable" if np.all(np.real(pde_eigenvals) < 0) else "Unstable"
            
            ws_pde.append([
                time, pde_max_eigen, pde_min_eigen, pde_cond, 
                pde_trace, pde_logdet, pde_stability
            ])
        except Exception as e:
            print(f"Error analyzing PDE Jacobian at time {time}: {str(e)}")
            ws_pde.append([time, "Error", "Error", "Error", "Error", "Error", "Error"])
        
        try:
            ws_details.cell(row=current_row, column=1, value=f"Time: {time}")
            current_row += 1
            
            ws_details.cell(row=current_row, column=1, value="ODE Jacobian Matrix:")
            current_row += 1
            for i, row in enumerate(ode_jac):
                for j, value in enumerate(row):
                    ws_details.cell(row=current_row+i, column=j+1, value=float(value))
            current_row += len(ode_jac) + 2
            
            ws_details.cell(row=current_row, column=1, value="PDE Jacobian Matrix:")
            current_row += 1
            for i, row in enumerate(pde_jac):
                for j, value in enumerate(row):
                    ws_details.cell(row=current_row+i, column=j+1, value=float(value))
            current_row += len(pde_jac) + 3
        except Exception as e:
            print(f"Error saving detailed matrices at time {time}: {str(e)}")
            current_row += 1
    
    try:
        wb.save(output_path)
        print(f"\nData saved to {output_path}")
    except Exception as e:
        print(f"Error saving workbook: {str(e)}")

def load_config(config_path: str) -> dict:
    """Load configuration from YAML file"""
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)

def main(args):
    # Load configuration
    config = load_config(args.config)
    
    # Set random seed
    set_random_seed(config['seed'])
    
    # Create output directory
    output_dir = Path(config['output_dir'])
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Create models
    density_model, price_model, optimizer, system_model, jacobian = create_models(
        params=config['model_params']
    )
    
    if args.mode == 'train':
        # Train agent
        env, agent = create_rl_system(
            system_model=system_model,
            price_model=price_model,
            history=None,
            env_config=config['environment'],
            agent_config=config['agent']
        )
        training_config = TrainingConfig(**config['training'])
        results = train_agent(env, agent, training_config)
        
    elif args.mode == 'evaluate':
        # Evaluate agent
        env, agent = create_rl_system(
            system_model=system_model,
            price_model=price_model,
            history=None,
            env_config=config['environment'],
            agent_config=config['agent']
        )
        results = evaluate_agent(env, agent, num_episodes=config['eval']['num_episodes'])
        
    elif args.mode == 'analyze':
        # System analysis
        bifurcation_params = BifurcationParameters(
            param_name=config['bifurcation']['param_name'],
            param_range=np.array(config['bifurcation']['param_range']),
            num_points=config['bifurcation'].get('num_points', DEFAULT_PARAMS['num_points']),
            max_iter=config['bifurcation'].get('max_iter', DEFAULT_PARAMS['max_iter']),
            tolerance=config['bifurcation'].get('tolerance', DEFAULT_PARAMS['tolerance'])
        )
        
        analyzer = BifurcationAnalyzer(system_model, bifurcation_params)
        
        # Execute dynamic system simulation
        q_rail_history, q_sea_history, price_rail_history, price_sea_history, \
        density_rail_history, density_sea_history, jacobian_history, simulation_time = \
            system_model.simulate_dynamic_system(config['simulation'])
        
        # Save time series data
        save_timestep_data_to_excel(
            str(output_dir / 'transport_chain_timesteps.xlsx'),
            system_model.rail_chains,
            system_model.sea_chains,
            q_rail_history,
            q_sea_history,
            price_rail_history,
            price_sea_history,
            simulation_time
        )
        
        # Analyze bifurcation
        bifurcation_results = analyzer.analyze_bifurcation(
            q_rail_history, q_sea_history, simulation_time
        )
        
        # Analyze Hopf bifurcation
        hopf_results = analyzer.analyze_hopf_bifurcation(
            q_rail_history, q_sea_history, simulation_time
        )
        
        # Plot bifurcation diagram
        analyzer.plot_bifurcation_diagram(
            q_rail_history, q_sea_history, simulation_time
        )
        
        # Analyze Jacobian matrix
        analyze_and_save_jacobian(
            jacobian_history,
            str(output_dir / 'jacobian_analysis.xlsx')
        )
        
        # Save analysis results
        results = {
            'bifurcation_results': bifurcation_results,
            'hopf_results': hopf_results
        }
        
    print(f"Results saved to {output_dir}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Transport Chain Control System")
    parser.add_argument('--config', type=str, default='config.yaml', 
                       help='Path to config file')
    parser.add_argument('--mode', type=str, 
                       choices=['train', 'evaluate', 'analyze'], 
                       default='train', 
                       help='Operation mode')
    args = parser.parse_args()
    main(args)